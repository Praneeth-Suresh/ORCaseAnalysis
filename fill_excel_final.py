"""
Fill the Excel answer sheet with the optimal LP solution.

Excel structure (Q1b sheet):
- Tool count section: rows 8-103
  - Each quarter block: 12 rows (6 mintech + 6 TOR)
  - Q1'26: rows 8-19 (FIXED, do not modify)
  - Q2'26: rows 20-31, Q3'26: rows 32-43, Q4'26: rows 44-55
  - Q1'27: rows 56-67, Q2'27: rows 68-79, Q3'27: rows 80-91, Q4'27: rows 92-103
  - Columns: C=Fab1, D=Fab2, E=Fab3

- Flow section: starts around row 8 (Q1'26 Node 1 Step 1)
  - Column O: quarter, P: node, Q: step, R: fab
  - Column S: loading (ANSWER CELL)
"""

import json
import openpyxl
import shutil

# Load solution
with open("results/q1b_correct_results.json") as f:
    result = json.load(f)

QUARTERS = ["Q1'26", "Q2'26", "Q3'26", "Q4'26", "Q1'27", "Q2'27", "Q3'27", "Q4'27"]
WS_ALL = ["A", "B", "C", "D", "E", "F", "A+", "B+", "C+", "D+", "E+", "F+"]
FABS = [1, 2, 3]
NODES = [1, 2, 3]
STEPS_COUNT = {1: 11, 2: 15, 3: 17}

tools_plan = [{ws: qd["tools"][ws] for ws in WS_ALL} for qd in result["tools_plan"]]
flow_plan = []
for qd in result["flow_plan"]:
    flow = {}
    for item in qd["flow"]:
        flow[(item["node"], item["step"] - 1, item["fab"])] = item["loading"]
    flow_plan.append(flow)

# Copy the original file
src = "5)BACC2026ParticipantAnswerSheet.xlsx"
dst = "results/BACC2026_AnswerSheet_v3.xlsx"
shutil.copy2(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb["Q1b"]

# ============================================================
# FILL TOOL COUNT SECTION
# ============================================================
# Q1'26 rows 8-19: FIXED, do not modify
# Q2'26 rows 20-31, Q3'26 rows 32-43, etc.

TOOL_ROW_START = {
    "Q1'26": 8,  # Fixed - do not modify
    "Q2'26": 20,
    "Q3'26": 32,
    "Q4'26": 44,
    "Q1'27": 56,
    "Q2'27": 68,
    "Q3'27": 80,
    "Q4'27": 92,
}

WS_ORDER = ["A", "B", "C", "D", "E", "F", "A+", "B+", "C+", "D+", "E+", "F+"]

filled_tool_cells = 0
for q_idx, quarter in enumerate(QUARTERS):
    if q_idx == 0:
        continue  # Q1'26 is fixed

    start_row = TOOL_ROW_START[quarter]
    tools = tools_plan[q_idx]

    for ws_idx, ws_name in enumerate(WS_ORDER):
        row = start_row + ws_idx
        fab1_val = tools[ws_name][0]
        fab2_val = tools[ws_name][1]
        fab3_val = tools[ws_name][2]

        ws.cell(row=row, column=3).value = fab1_val  # Fab 1
        ws.cell(row=row, column=4).value = fab2_val  # Fab 2
        ws.cell(row=row, column=5).value = fab3_val  # Fab 3
        filled_tool_cells += 3

print(f"Filled {filled_tool_cells} tool count cells")

# ============================================================
# FILL FLOW SECTION
# ============================================================
# Find the flow section by scanning for the answer column (S = col 19)
# The flow section has: col O=quarter, col P=node, col Q=step, col R=fab, col S=loading

# First, find the start row of the flow section
flow_start_row = None
for row in range(1, 200):
    cell_o = ws.cell(row=row, column=15)  # Column O
    cell_p = ws.cell(row=row, column=16)  # Column P
    if cell_o.value == "Q1'26" and cell_p.value == 1:
        flow_start_row = row
        break

if flow_start_row is None:
    # Try to find it by scanning
    for row in range(1, 300):
        cell_o = ws.cell(row=row, column=15)
        if cell_o.value is not None and "Q1" in str(cell_o.value):
            print(f"Found potential flow start at row {row}: {cell_o.value}")
            flow_start_row = row
            break

print(f"Flow section starts at row: {flow_start_row}")

# Scan the flow section and fill column S
# The flow section rows overlap with the tool count section
# Column O (15) = quarter, P (16) = node, Q (17) = step, R (18) = fab
# Column S (19) = loading (ANSWER CELL)
filled_flow_cells = 0
max_row = ws.max_row

for row in range(1, max_row + 1):
    quarter_val = ws.cell(row=row, column=15).value  # Col O
    node_val = ws.cell(row=row, column=16).value  # Col P
    step_val = ws.cell(row=row, column=17).value  # Col Q
    fab_val = ws.cell(row=row, column=18).value  # Col R

    if quarter_val is None or node_val is None:
        continue

    # Normalize quarter string (handle special apostrophe characters)
    quarter_str = str(quarter_val).strip()
    # Normalize apostrophes
    quarter_str = quarter_str.replace("\u2019", "'").replace("\u2018", "'")
    if quarter_str not in QUARTERS:
        continue

    try:
        node = int(node_val)
        step = int(step_val) - 1  # Convert to 0-indexed
        fab = int(fab_val)
    except (ValueError, TypeError):
        continue

    q_idx = QUARTERS.index(quarter_str)
    flow = flow_plan[q_idx]
    loading = flow.get((node, step, fab), 0)

    ws.cell(row=row, column=19).value = round(loading)  # Col S
    filled_flow_cells += 1

print(f"Filled {filled_flow_cells} flow cells")

# ============================================================
# SAVE
# ============================================================
wb.save(dst)
print(f"\nSaved: {dst}")
print(f"Total cells filled: {filled_tool_cells + filled_flow_cells}")

# ============================================================
# VERIFY KEY CELLS
# ============================================================
wb2 = openpyxl.load_workbook(dst)
ws2 = wb2["Q1b"]

print("\nVerification - Q1'26 tool counts (should be unchanged):")
for row in range(8, 20):
    ws_name = ws2.cell(row=row, column=2).value
    f1 = ws2.cell(row=row, column=3).value
    f2 = ws2.cell(row=row, column=4).value
    f3 = ws2.cell(row=row, column=5).value
    print(f"  {ws_name}: F1={f1}, F2={f2}, F3={f3}")

print("\nQ2'26 tool counts (first few):")
for row in range(20, 32):
    ws_name = ws2.cell(row=row, column=2).value
    f1 = ws2.cell(row=row, column=3).value
    f2 = ws2.cell(row=row, column=4).value
    f3 = ws2.cell(row=row, column=5).value
    print(f"  {ws_name}: F1={f1}, F2={f2}, F3={f3}")

print("\nFlow section sample (first 9 rows):")
for row in range(flow_start_row, flow_start_row + 9):
    q = ws2.cell(row=row, column=15).value
    n = ws2.cell(row=row, column=16).value
    s = ws2.cell(row=row, column=17).value
    f = ws2.cell(row=row, column=18).value
    l = ws2.cell(row=row, column=19).value
    print(f"  R{row}: Q={q}, N={n}, S={s}, F={f}, Loading={l}")

wb2.close()
