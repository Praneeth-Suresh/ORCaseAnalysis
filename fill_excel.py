"""
Script 8: Fill Excel Answer Sheet (Part A & Part B)
===================================================
Fills the Q1a or Q1b answer sheet using the corresponding results JSON.
Prompts the user to choose which part to fill (A or B). Part A pulls from
results/q1a_results.json; Part B pulls from results/q1b_solution_v2_results.json.

Answer areas:
1. Tool count section: cols C, D, E (Fab1, Fab2, Fab3 WS counts) for rows 8-103
    - Each quarter has 12 rows: 6 mintech (A-F) + 6 TOR (A+-F+)
    - Q1'26: rows 8-19, Q2'26: rows 20-31, Q3'26: rows 32-43, Q4'26: rows 44-55
    - Q1'27: rows 56-67, Q2'27: rows 68-79, Q3'27: rows 80-91, Q4'27: rows 92-103

2. Flow section: col S for rows 8-1039
    - Each row: (quarter, node, step, fab) -> loading value
    - Part A: aggregate per-(quarter,node,fab) from first step (mintech+TOR)
    - Part B: use assignment per (quarter,node,fab) directly from the Part B JSON
"""

import json
import math
import shutil
from pathlib import Path

import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# PATHS & LOAD SOLUTIONS
# ─────────────────────────────────────────────────────────────────────────────

ROOT = Path(__file__).resolve().parent
RESULTS_DIR = ROOT / "results"
SRC_SHEET = ROOT / "5)BACC2026ParticipantAnswerSheet.xlsx"
DST_SHEET = RESULTS_DIR / "BACC2026_FilledAnswerSheet_DP.xlsx"
Q1A_RESULTS_PATH = RESULTS_DIR / "q1a_results.json"
Q1B_RESULTS_PATH = (
    RESULTS_DIR / "dp_lean_results.json"
)  # DP optimal (gran=1000, $5,009,100,000)
PARAMS_PATH = ROOT / "parameters" / "params.json"

with open(Q1A_RESULTS_PATH) as f:
    SOL_A = json.load(f)

with open(Q1B_RESULTS_PATH) as f:
    _dp_raw = json.load(f)

with open(PARAMS_PATH) as f:
    _P = json.load(f)

# ───────────────────────────────────────────────────────────────────────────────
# BUILD SOL_B FROM DP PATH
# Converts dp_results.json (list-of-quarters format) into the dict format
# expected by get_tool_count_b / get_flow_b:
#   assignment : {quarter: {node_str: {fab_str: wafers}}}
#   tor_tools  : {quarter: {ws: {fab_str: count}}}   (cumulative — never decreases)
#   mt_tools   : {quarter: {ws: {fab_str: count}}}   (after greedy move-outs)
# ───────────────────────────────────────────────────────────────────────────────


def _build_sol_b(dp_raw, params):
    """Derive assignment, tor_tools, and mt_tools from the DP optimal path."""
    _fabs = [1, 2, 3]
    _nodes = [1, 2, 3]
    _mt_ws = ["A", "B", "C", "D", "E", "F"]
    _tor_ws = ["A+", "B+", "C+", "D+", "E+", "F+"]

    _proc = {
        int(n): [(s["ws_tor"], s["rpt_tor"]) for s in steps]
        for n, steps in params["process_steps"].items()
    }
    _ws_specs = {
        ws: (d["space_m2"], d["utilization"]) for ws, d in params["ws_specs"].items()
    }
    _min_wk = params["minutes_per_week"]
    _moveout_cost = params["moveout_cost_per_tool"]
    _fab_space = {int(f): d["space"] for f, d in params["fab_specs"].items()}
    _init_mt = {
        ws: {int(f): d["tools"][ws] for f, d in params["fab_specs"].items()}
        for ws in _mt_ws
    }
    _mt_space = {ws: params["ws_specs"][ws]["space_m2"] for ws in _mt_ws}
    _tor_space = {ws: params["ws_specs"][ws]["space_m2"] for ws in _tor_ws}

    # TOR tools needed per wafer of each node
    _tor_per_wfr = {}
    for n in _nodes:
        _tor_per_wfr[n] = {ws: 0.0 for ws in _tor_ws}
        for ws_tor, rpt_tor in _proc[n]:
            _, util = _ws_specs[ws_tor]
            _tor_per_wfr[n][ws_tor] += rpt_tor / (_min_wk * util)

    assignment = {}
    tor_tools = {}
    mt_tools = {}

    # Mutable inventory state
    cur_mt = {ws: {f: _init_mt[ws][f] for f in _fabs} for ws in _mt_ws}
    cur_tor = {ws: {f: 0 for f in _fabs} for ws in _tor_ws}

    for q_entry in dp_raw["dp_path"]:
        q = q_entry["quarter"]
        asgn = q_entry["assignment"]  # {node_str: {fab_str: wafers}}

        # ── assignment ──
        assignment[q] = {
            str(n): {str(f): int(asgn[str(n)][str(f)]) for f in _fabs} for n in _nodes
        }

        # ── TOR tools needed this quarter (continuous, then ceiling-rounded) ──
        tor_cont = {ws: {f: 0.0 for f in _fabs} for ws in _tor_ws}
        for f in _fabs:
            for n in _nodes:
                wafers = asgn[str(n)][str(f)]
                if wafers > 0:
                    for ws in _tor_ws:
                        tor_cont[ws][f] += wafers * _tor_per_wfr[n][ws]
        tor_req = {ws: {f: math.ceil(tor_cont[ws][f]) for f in _fabs} for ws in _tor_ws}

        # TOR inventory is cumulative (tools cannot be removed)
        for ws in _tor_ws:
            for f in _fabs:
                cur_tor[ws][f] = max(cur_tor[ws][f], tor_req[ws][f])

        # TOR space needed per fab based on current inventory
        tor_sp = {
            f: sum(_tor_space[ws] * cur_tor[ws][f] for ws in _tor_ws) for f in _fabs
        }

        # ── Greedy MT move-out (largest footprint first) ──
        for f in _fabs:
            mt_sp = sum(_mt_space[ws] * cur_mt[ws][f] for ws in _mt_ws)
            excess = tor_sp[f] + mt_sp - _fab_space[f]
            if excess > 0.001:
                for ws in sorted(_mt_ws, key=lambda w: -_mt_space[w]):
                    if excess <= 0.001:
                        break
                    avail = cur_mt[ws][f]
                    if avail <= 0:
                        continue
                    to_move = min(avail, math.ceil(excess / _mt_space[ws]))
                    cur_mt[ws][f] -= to_move
                    excess -= to_move * _mt_space[ws]

        # ── Store tool snapshots ──
        tor_tools[q] = {ws: {str(f): cur_tor[ws][f] for f in _fabs} for ws in _tor_ws}
        mt_tools[q] = {ws: {str(f): cur_mt[ws][f] for f in _fabs} for ws in _mt_ws}

    return {"assignment": assignment, "tor_tools": tor_tools, "mt_tools": mt_tools}


SOL_B = _build_sol_b(_dp_raw, _P)
print(f"Part B optimal cost (DP gran=1000): ${_dp_raw['dp_coarse_cost']:,.0f}")

QUARTERS = ["Q1'26", "Q2'26", "Q3'26", "Q4'26", "Q1'27", "Q2'27", "Q3'27", "Q4'27"]
FABS = [1, 2, 3]
NODES = [1, 2, 3]
MINTECH_WS = ["A", "B", "C", "D", "E", "F"]
TOR_WS = ["A+", "B+", "C+", "D+", "E+", "F+"]
ALL_WS = MINTECH_WS + TOR_WS


def safe_get(dic, *keys, default=0):
    cur = dic
    for k in keys:
        if cur is None:
            return default
        cur = cur.get(k)
    return default if cur is None else cur


# Tool counts: {quarter: {ws: {fab: count}}}
def get_tool_count_a(q, ws, f):
    q = normalize_q(q)
    return safe_get(SOL_A, "total_tools", q, ws, str(f), default=0)


def get_tool_count_b(q, ws, f):
    q = normalize_q(q)
    if ws in TOR_WS:
        return safe_get(SOL_B, "tor_tools", q, ws, str(f), default=0)
    return safe_get(SOL_B, "mt_tools", q, ws, str(f), default=0)


# Quarter key normalization: Excel uses Unicode right single quote \u2019, JSON uses ASCII apostrophe
def normalize_q(q):
    return q.replace("\u2019", "'").replace("\u2018", "'")


# Flow (aggregate per quarter/node/fab) for Part A: use first step's mintech+TOR flow
def get_flow_a(q, n, f):
    q_norm = normalize_q(q)
    flow_mt_q = safe_get(SOL_A, "flow_mt", q_norm, str(n), default={})
    flow_tor_q = safe_get(SOL_A, "flow_tor", q_norm, str(n), default={})

    def first_step(flow_dict):
        if not flow_dict:
            return None
        try:
            return min(flow_dict.keys(), key=lambda s: int(s))
        except Exception:
            return next(iter(flow_dict.keys()))

    step_key = first_step(flow_mt_q) or first_step(flow_tor_q)
    if step_key is None:
        return 0

    mt_val = safe_get(flow_mt_q, step_key, str(f), default=0)
    tor_val = safe_get(flow_tor_q, step_key, str(f), default=0)
    return mt_val + tor_val


# Flow for Part B: assignment per (quarter,node,fab)
def get_flow_b(q, n, f):
    q_norm = normalize_q(q)
    return safe_get(SOL_B, "assignment", q_norm, str(n), str(f), default=0)


# ─────────────────────────────────────────────────────────────────────────────
# FILL SHEET (generic)
# ─────────────────────────────────────────────────────────────────────────────


def fill_sheet(wb, sheet_name, part):
    ws = wb[sheet_name]

    # ── Tool count section ──────────────────────────────────────────────────
    # Row mapping: Q1'26 starts at row 8, each quarter has 12 rows (6 MT + 6 TOR)
    QUARTER_START_ROWS = {
        "Q1'26": 8,
        "Q2'26": 20,
        "Q3'26": 32,
        "Q4'26": 44,
        "Q1'27": 56,
        "Q2'27": 68,
        "Q3'27": 80,
        "Q4'27": 92,
    }

    tools_filled = 0
    for q, start_row in QUARTER_START_ROWS.items():
        for i, tool_ws in enumerate(ALL_WS):
            row = start_row + i
            for col_idx, fab in [(3, 1), (4, 2), (5, 3)]:  # C=Fab1, D=Fab2, E=Fab3
                if part == "A":
                    count = get_tool_count_a(q, tool_ws, fab)
                else:
                    count = get_tool_count_b(q, tool_ws, fab)
                ws.cell(row, col_idx).value = count
                tools_filled += 1

    print(f"  Filled {tools_filled} tool count cells")

    # ── Flow section ────────────────────────────────────────────────────────
    # Col S (19) = Loading (to fill)
    flow_filled = 0
    for r in range(8, 1040):
        q_val = ws.cell(r, 15).value  # col O = Quarter
        n_val = ws.cell(r, 16).value  # col P = Node
        f_val = ws.cell(r, 18).value  # col R = Fab

        if q_val is not None and n_val is not None and f_val is not None:
            if part == "A":
                flow = get_flow_a(q_val, int(n_val), int(f_val))
            else:
                flow = get_flow_b(q_val, int(n_val), int(f_val))
            ws.cell(r, 19).value = flow
            flow_filled += 1

    print(f"  Filled {flow_filled} flow cells")

    return tools_filled + flow_filled


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    choice = input("Which part to fill? Enter A or B: ").strip().upper()
    if choice not in {"A", "B"}:
        raise SystemExit("Invalid choice. Please enter 'A' or 'B'.")

    sheet_name = "Q1a" if choice == "A" else "Q1b"

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SRC_SHEET, DST_SHEET)
    print(f"Copied answer sheet to: {DST_SHEET}")

    wb = openpyxl.load_workbook(DST_SHEET)
    print(f"Sheets: {wb.sheetnames}")

    if sheet_name in wb.sheetnames:
        print(f"\nFilling {sheet_name} sheet (Part {choice} results)...")
        n = fill_sheet(wb, sheet_name, choice)
        print(f"  Total cells filled: {n}")
    else:
        raise SystemExit(f"Sheet {sheet_name} not found in workbook.")

    wb.save(DST_SHEET)
    print(f"\nSaved filled answer sheet to: {DST_SHEET}")

    # Verify by reading back (first block only)
    print("\n=== Verification ===")
    wb2 = openpyxl.load_workbook(DST_SHEET)
    ws2 = wb2[sheet_name]
    print(f"\n{sheet_name} - Tool counts (Q1'26, rows 8-19):")
    for r in range(8, 20):
        b = ws2.cell(r, 2).value  # WS name
        c = ws2.cell(r, 3).value  # Fab 1
        d = ws2.cell(r, 4).value  # Fab 2
        e = ws2.cell(r, 5).value  # Fab 3
        print(f"  Row {r}: {b} | F1={c}, F2={d}, F3={e}")

    print(f"\n{sheet_name} - Flow (first 10 rows):")
    for r in range(8, 18):
        q = ws2.cell(r, 15).value
        n = ws2.cell(r, 16).value
        s = ws2.cell(r, 17).value
        f = ws2.cell(r, 18).value
        loading = ws2.cell(r, 19).value
        print(f"  Row {r}: {q} Node{n} Step{s} Fab{f} -> {loading}")
